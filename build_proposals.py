#!/usr/bin/env python3
"""把「2B提案簡報檢索表」xlsx 的兩張歷年案例表合併成 proposals.js 快照。"""
import openpyxl, json, os, re

XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_index.xlsx")
OUT  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "proposals.js")

wb = openpyxl.load_workbook(XLSX)  # 保留 hyperlink

def cell_val(ws, r, c):
    v = ws.cell(r, c).value
    if v is None: return ""
    return str(v).strip()

def cell_link(ws, r, c):
    h = ws.cell(r, c).hyperlink
    return h.target if h else None

# 兩張表的元件欄位對照（欄號: 標籤）
SHEETS = {
    "B. 歷年案例": {
        "data_start": 3, "name_col": 9, "price_col": 10, "note_col": 11,
        "meal_col": None, "issue_col": None,
        "comp": {12:"導覽",13:"實境",14:"其他遊戲",15:"手作體驗",16:"空間參訪",17:"餐食",
                 18:"遊船",19:"伴手禮",20:"會議空間",21:"交通",22:"住宿",23:"其他"},
    },
    "專案企劃歷屆案例整理": {
        "data_start": 3, "name_col": 9, "price_col": 11, "note_col": 17,
        "meal_col": 12, "issue_col": 13,
        "comp": {18:"導覽",19:"議題",20:"Teambuilding",21:"其他遊戲",22:"志工服務",23:"手作體驗",
                 24:"空間參訪",25:"餐食",26:"遊船",27:"伴手禮",28:"會議空間",29:"交通",30:"住宿",
                 31:"其他",32:"新路線開發",33:"路線上架",34:"客服服務",35:"社群宣傳",36:"店家串聯"},
    },
}

def parse_name(name):
    """命名規則：年月日_企業_負責人_備註。抽出企業名與日期（盡量）。"""
    parts = re.split(r"[_＿]", name)
    date = company = owner = ""
    if parts:
        p0 = parts[0].strip()
        if re.fullmatch(r"\d{6}|\d{8}|\d{4}|20\d{2}\d{0,4}", p0.replace(".","")):
            date = p0
            company = parts[1].strip() if len(parts) > 1 else ""
            owner = parts[2].strip() if len(parts) > 2 else ""
        else:
            company = p0
            owner = parts[2].strip() if len(parts) > 2 else ""
    return date, company, owner

out = []
for sheet, cfg in SHEETS.items():
    ws = wb[sheet]
    for r in range(cfg["data_start"], ws.max_row + 1):
        name = cell_val(ws, r, cfg["name_col"])
        if not name or name in ("案例","0") or len(name) < 2:
            continue
        if name.startswith("命名") or "Sales Kit" in name:
            continue
        city = cell_val(ws, r, 2)
        spots = [cell_val(ws, r, c) for c in (3,4,5)]
        spots = [s for s in spots if s]
        duration = cell_val(ws, r, 6)
        headcount = cell_val(ws, r, 7)
        purpose = cell_val(ws, r, 8)
        price = cell_val(ws, r, cfg["price_col"])
        note = cell_val(ws, r, cfg["note_col"])
        link = cell_link(ws, r, cfg["name_col"])
        meal = cell_val(ws, r, cfg["meal_col"]) if cfg["meal_col"] else ""
        issue = cell_val(ws, r, cfg["issue_col"]) if cfg["issue_col"] else ""
        comps, comp_detail = [], []
        for col, label in cfg["comp"].items():
            v = cell_val(ws, r, col)
            if v and v != "0":
                comps.append(label)
                if not re.fullmatch(r"\d+", v):
                    comp_detail.append(f"{label}:{v}")
        date, company, owner = parse_name(name)
        out.append({
            "name": name, "link": link, "company": company or name, "date": date, "owner": owner,
            "city": city, "spots": spots, "duration": duration, "headcount": headcount,
            "purpose": purpose, "pricePP": price, "mealTransport": meal, "issue": issue,
            "components": comps, "compDetail": comp_detail, "note": note, "source": sheet,
        })

# 排序：有日期者新到舊，其餘殿後
def k(p):
    d = re.sub(r"\D","",p["date"])
    return (0 if d else 1, -(int(d) if d.isdigit() else 0))
out.sort(key=k)

with open(OUT, "w", encoding="utf-8") as f:
    f.write("// 過去提案快照（取自『2B提案簡報檢索表』兩張歷年案例表，point-in-time 匯出）\n")
    f.write("// 由 build_proposals.py 產生。\n")
    f.write("window.PROPOSALS = ")
    json.dump(out, f, ensure_ascii=False, indent=0)
    f.write(";\n")

from collections import Counter
withlink = sum(1 for p in out if p["link"])
print(f"提案筆數：{len(out)}（含簡報連結 {withlink}）")
print("來源分頁：", dict(Counter(p["source"] for p in out)))
print("活動目的 Top:", dict(Counter(p["purpose"] for p in out if p["purpose"]).most_common(8)))
print("地區 Top:", dict(Counter(p["city"] for p in out if p["city"]).most_common(8)))
print("範例：")
for p in out[:3]:
    print("  ", p["company"], "|", p["city"], "/".join(p["spots"]), "|", p["duration"], p["headcount"], "|", p["purpose"], "|/pp", p["pricePP"], "|", "連結✓" if p["link"] else "無連結")
print("輸出：", OUT)
